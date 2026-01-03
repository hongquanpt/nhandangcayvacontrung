# -*- coding: utf-8 -*-
"""
CuteID - Nhận diện côn trùng & thực vật (NO BBOX)
- Upload ảnh / chụp ảnh / Realtime (WebRTC)
- Hiển thị kết quả trong popup (dialog/modal)
- Đọc Wikipedia tiếng Việt (VI) bằng giọng Google nữ ngay sau khi nhận diện

Cài đặt:
  pip install streamlit opencv-python numpy requests openpyxl streamlit-webrtc streamlit-autorefresh gTTS

Chạy:
  streamlit run app.py

Khuyến nghị: KHÔNG hard-code API key trong code.
  Windows (PowerShell):
    setx PLANTNET_API_KEY "xxx"
    setx INSECT_ID_API_KEY "yyy"
  macOS/Linux:
    export PLANTNET_API_KEY="xxx"
    export INSECT_ID_API_KEY="yyy"
"""

import os
import json
import time
import base64
import re
from io import BytesIO
import threading
import random
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Any, Optional, Tuple

import numpy as np
import cv2
import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from streamlit_webrtc import webrtc_streamer, VideoTransformerBase, WebRtcMode

# (Tuỳ chọn) auto-refresh để UI “bắt” kết quả mới và đọc gần như ngay lập tức
try:
    from streamlit_autorefresh import st_autorefresh  # pip install streamlit-autorefresh
except Exception:
    st_autorefresh = None


# =========================================================
# CONFIG
# =========================================================
APP_NAME = "CuteID 🪲🌿 — Nhận diện côn trùng & thực vật (NO BBOX)"

# Lấy API key từ env (không hard-code)
PLANTNET_API_KEY = "2b10h0g5knI9zkfmSeANOfjkPu"
INSECT_ID_API_KEY = "mGqkBybFTtcovHBhfI6quqEJPHrpCGTwjYvV6yTFFSL1Q6Hjhn"
WIKI_USER_AGENT_DEFAULT = "SpeciesReader/1.0 (contact: nguyenhongquan122002@gmail.com)" 

DATA_ROOT = "data"
COLLECTION_DIR = os.path.join(DATA_ROOT, "Bo_suu_tap")
UNKNOWN_DIR = os.path.join(DATA_ROOT, "Chua_train")
LOG_XLSX = os.path.join(DATA_ROOT, "logs.xlsx")
WIKI_CACHE = os.path.join(DATA_ROOT, "wiki_cache.json")

# Kindwise/InsectID normalize config (an toàn, giảm lỗi 400 do ảnh quá lớn/PNG)
KINDWISE_MAX_SIDE = 1500
KINDWISE_MAX_PIXELS = 25_000_000
KINDWISE_JPEG_QUALITY = 85
KINDWISE_TIMEOUT = 30

JSON_LOCK = threading.Lock()
LOG_LOCK = threading.Lock()


# =========================================================
# UI THEME (cute & modern)
# =========================================================
def apply_cute_theme() -> None:
    css = r"""
    <style>
      /* App background */
      .stApp {
        background: radial-gradient(1200px 800px at 10% 10%, rgba(255, 209, 220, 0.35), transparent 55%),
                    radial-gradient(900px 700px at 90% 20%, rgba(186, 230, 253, 0.40), transparent 55%),
                    radial-gradient(900px 700px at 30% 90%, rgba(187, 247, 208, 0.35), transparent 55%),
                    linear-gradient(180deg, #ffffff 0%, #fbfbff 40%, #ffffff 100%);
      }

      /* Hide Streamlit default decorations a bit */
      #MainMenu {visibility: hidden;}
      footer {visibility: hidden;}
      header {visibility: hidden;}

      /* Hero */
      .cute-hero {
        padding: 18px 18px;
        border-radius: 18px;
        background: rgba(255,255,255,0.72);
        border: 1px solid rgba(0,0,0,0.05);
        box-shadow: 0 14px 40px rgba(0,0,0,0.08);
        backdrop-filter: blur(10px);
        margin-bottom: 14px;
      }
      .cute-hero h1 {
        margin: 0;
        font-size: 30px;
        line-height: 1.2;
      }
      .cute-hero p {
        margin: 8px 0 0;
        color: rgba(0,0,0,0.62);
        font-size: 14px;
      }

      /* Cards */
      .cute-card {
        padding: 14px 14px;
        border-radius: 16px;
        background: rgba(255,255,255,0.78);
        border: 1px solid rgba(0,0,0,0.05);
        box-shadow: 0 12px 34px rgba(0,0,0,0.06);
        backdrop-filter: blur(10px);
      }

      /* Buttons (primary look) */
      div.stButton > button {
        border-radius: 14px !important;
        padding: 0.55rem 1.0rem !important;
        border: 1px solid rgba(0,0,0,0.08) !important;
        box-shadow: 0 10px 24px rgba(0,0,0,0.10) !important;
      }
      div.stButton > button:hover {
        transform: translateY(-1px);
        transition: 0.15s ease;
      }

      /* Inputs */
      .stTextInput > div > div > input,
      .stSelectbox > div > div > div,
      .stSlider > div {
        border-radius: 14px !important;
      }

      /* Badge */
      .badge {
        display: inline-block;
        padding: 4px 10px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 700;
        border: 1px solid rgba(0,0,0,0.08);
      }
      .badge-known { background: rgba(187, 247, 208, 0.65); }
      .badge-unknown { background: rgba(254, 202, 202, 0.65); }
      .badge-info { background: rgba(186, 230, 253, 0.65); }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)


def badge_html(text: str, kind: str = "info") -> str:
    kind = (kind or "info").lower()
    klass = "badge-info"
    if kind == "known":
        klass = "badge-known"
    elif kind == "unknown":
        klass = "badge-unknown"
    return f"<span class='badge {klass}'>{text}</span>"


def card_open() -> None:
    st.markdown("<div class='cute-card'>", unsafe_allow_html=True)


def card_close() -> None:
    st.markdown("</div>", unsafe_allow_html=True)


# =========================================================
# FS + EXCEL
# =========================================================
def ensure_dirs() -> None:
    os.makedirs(DATA_ROOT, exist_ok=True)
    os.makedirs(COLLECTION_DIR, exist_ok=True)
    os.makedirs(UNKNOWN_DIR, exist_ok=True)


def init_log() -> None:
    with LOG_LOCK:
        if not os.path.exists(LOG_XLSX):
            wb = Workbook()
            ws = wb.active
            ws.title = "log"
            ws.append(
                [
                    "timestamp",
                    "mode",
                    "source",
                    "kingdom_guess",
                    "label_scientific",
                    "label_common",
                    "confidence",
                    "status",
                    "image_path",
                    "wiki_title_query",
                    "wiki_used_lang",
                    "wiki_title_used",
                    "wiki_intro_vi",
                    "wiki_intro_en",
                ]
            )
            wb.save(LOG_XLSX)


def append_log(row: Dict[str, Any]) -> None:
    with LOG_LOCK:
        wb = load_workbook(LOG_XLSX)
        ws = wb["log"]
        ws.append(
            [
                row.get("timestamp", ""),
                row.get("mode", ""),
                row.get("source", ""),
                row.get("kingdom_guess", ""),
                row.get("label_scientific", ""),
                row.get("label_common", ""),
                float(row.get("confidence", 0.0)),
                row.get("status", ""),
                row.get("image_path", ""),
                row.get("wiki_title_query", ""),
                row.get("wiki_used_lang", ""),
                row.get("wiki_title_used", ""),
                row.get("wiki_intro_vi", ""),
                row.get("wiki_intro_en", ""),
            ]
        )
        wb.save(LOG_XLSX)


def safe_filename(s: str) -> str:
    s = (s or "").strip()
    out = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        elif ch.isspace():
            out.append("_")
        else:
            out.append("_")
    return "".join(out).strip().replace(" ", "_")[:120] or "unknown"


def save_jpeg_bytes(jpeg_bytes: bytes, out_dir: str, label: str, conf: float) -> str:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fn = f"{ts}_{safe_filename(label)}_{conf:.2f}.jpg"
    path = os.path.join(out_dir, fn)
    with open(path, "wb") as f:
        f.write(jpeg_bytes)
    return path


def load_json(path: str, default: Any) -> Any:
    with JSON_LOCK:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return default
        return default


def save_json(path: str, data: Any) -> None:
    tmp = path + ".tmp"
    with JSON_LOCK:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)


# =========================================================
# IMAGE UTILS (fix PNG upload -> JPEG thật, resize cho InsectID)
# =========================================================
def decode_image_to_bgr(image_bytes: bytes) -> Optional[np.ndarray]:
    if not image_bytes:
        return None
    npimg = np.frombuffer(image_bytes, np.uint8)
    return cv2.imdecode(npimg, cv2.IMREAD_COLOR)


def encode_bgr_to_jpeg_bytes(bgr: np.ndarray, quality: int = 92) -> bytes:
    quality = int(max(30, min(95, quality)))
    ok, buf = cv2.imencode(".jpg", bgr, [int(cv2.IMWRITE_JPEG_QUALITY), quality])
    return buf.tobytes() if ok else b""


def normalize_for_kindwise(image_bytes: bytes) -> Tuple[bytes, Dict[str, Any]]:
    """
    Ép JPEG + resize để tránh 400 (ảnh quá lớn/PNG/định dạng lạ).
    """
    bgr = decode_image_to_bgr(image_bytes)
    if bgr is None:
        return b"", {"error": "Không decode được ảnh để normalize."}

    h, w = bgr.shape[:2]
    orig_pixels = int(w) * int(h)

    scale = 1.0
    if max(w, h) > KINDWISE_MAX_SIDE:
        scale = min(scale, KINDWISE_MAX_SIDE / float(max(w, h)))
    if orig_pixels > KINDWISE_MAX_PIXELS:
        scale = min(scale, (KINDWISE_MAX_PIXELS / float(orig_pixels)) ** 0.5)

    new_w, new_h = w, h
    resized = bgr
    if scale < 1.0:
        new_w = max(1, int(round(w * scale)))
        new_h = max(1, int(round(h * scale)))
        resized = cv2.resize(bgr, (new_w, new_h), interpolation=cv2.INTER_AREA)

    jpeg_bytes = encode_bgr_to_jpeg_bytes(resized, quality=KINDWISE_JPEG_QUALITY)
    if not jpeg_bytes:
        return b"", {"error": "Encode JPEG thất bại khi normalize."}

    info = {
        "orig_wh": [w, h],
        "new_wh": [new_w, new_h],
        "orig_pixels": orig_pixels,
        "new_pixels": int(new_w) * int(new_h),
        "jpeg_len": len(jpeg_bytes),
        "max_side": KINDWISE_MAX_SIDE,
        "max_pixels": KINDWISE_MAX_PIXELS,
        "jpeg_quality": KINDWISE_JPEG_QUALITY,
    }
    return jpeg_bytes, info


def _resp_debug(r: requests.Response, limit: int = 2000) -> Dict[str, Any]:
    ct = (r.headers.get("Content-Type") or "").lower()
    out: Dict[str, Any] = {
        "status_code": r.status_code,
        "content_type": ct,
        "headers_subset": {k: r.headers.get(k) for k in ["Date", "Content-Type", "x-request-id", "cf-ray"] if r.headers.get(k)},
        "text_snippet": "",
        "json": None,
    }
    try:
        if "application/json" in ct:
            out["json"] = r.json()
        else:
            out["text_snippet"] = (r.text or "")[:limit]
    except Exception:
        out["text_snippet"] = (r.text or "")[:limit]
    return out


# =========================================================
# TTS (Trình duyệt) - fallback khi chưa cài gTTS
# =========================================================
def browser_tts_sequence(items):
    """
    items: [{"text": "...", "lang": "vi-VN", "prefer_google": true/false}, ...]
    Đọc lần lượt theo thứ tự items (không bị cancel giữa các đoạn).
    - Nếu prefer_google=True: ưu tiên chọn voice có chữ "Google" (thường là giọng Google trên Chrome).
    """
    items = items or []
    items = [x for x in items if (x.get("text") or "").strip()]
    if not items:
        return

    items_json = json.dumps(items, ensure_ascii=False)

    html = f"""
    <script>
      (function() {{
        const synth = window.speechSynthesis;
        if (!synth) return;

        const items = {items_json};

        function pickVoice(u, preferGoogle) {{
          const voices = synth.getVoices() || [];
          const want = (u.lang || "").toLowerCase().split("-")[0]; // vi / en
          const isGoogle = (v) => ((v.name || "").toLowerCase().includes("google"));
          const langOk = (v) => ((v.lang || "").toLowerCase().startsWith(want));

          let v = null;

          if (preferGoogle) {{
            v = voices.find(x => langOk(x) && isGoogle(x));
          }}
          if (!v) {{
            v = voices.find(x => langOk(x));
          }}
          if (!v && want === "vi") {{
            v = voices.find(x => isGoogle(x) && ((x.lang || "").toLowerCase().includes("vi")));
          }}
          if (!v) {{
            v = voices[0] || null;
          }}
          if (v) u.voice = v;
        }}

        function run() {{
          synth.cancel();

          const utterances = items.map(it => {{
            const u = new SpeechSynthesisUtterance(it.text);
            u.lang = it.lang || "vi-VN";
            u.pitch = (it.pitch !== undefined) ? it.pitch : 1.06;
            u.rate  = (it.rate  !== undefined) ? it.rate  : 1.00;
            pickVoice(u, !!it.prefer_google);
            return u;
          }});

          function speakIdx(i) {{
            if (i >= utterances.length) return;
            const u = utterances[i];
            u.onend = () => speakIdx(i + 1);
            synth.speak(u);
          }}

          speakIdx(0);
        }}

        const voicesNow = synth.getVoices();
        if (voicesNow && voicesNow.length) {{
          run();
        }} else {{
          synth.onvoiceschanged = () => {{
            run();
            synth.onvoiceschanged = null;
          }};
        }}
      }})();
    </script>
    """
    st.components.v1.html(html, height=0)

def make_speech_items(scientific: str, common: str, status: str, read_vi: bool, read_en: bool):
    scientific = (scientific or "").strip()
    common = (common or "").strip()
    status = (status or "").strip().upper()

    items = []

    if status == "UNKNOWN":
        if read_vi:
            items.append({"text": "Chưa nhận diện chắc chắn. Đã lưu vào thư mục chưa train.", "lang": "vi-VN"})
        if read_en:
            items.append({"text": "Not confident. Saved to the untrained folder.", "lang": "en-US"})
        return items


# =========================================================
# TTS: Đọc Wikipedia (VI) bằng giọng Google (gTTS)
# =========================================================
# Yêu cầu: pip install gTTS
try:
    from gtts import gTTS  # type: ignore
except Exception:
    gTTS = None


def _clean_wiki_vi_for_tts(text: str, max_chars: int = 900) -> str:
    """Làm sạch text wiki để đọc dễ nghe + tránh quá dài."""
    t = (text or "").strip()
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"\[[0-9]+\]", "", t)          # bỏ citation kiểu [1], [2]...
    t = re.sub(r"\s*\([^)]*\)", "", t)       # bỏ ngoặc tròn (phiên âm/dẫn chiếu)
    t = t.strip()

    if len(t) > max_chars:
        cut = t[:max_chars]
        idx = max(cut.rfind("."), cut.rfind("!"), cut.rfind("?"))
        if idx >= 200:
            t = cut[: idx + 1]
        else:
            t = cut
    return t


def _gtts_vi_mp3(text: str) -> Optional[bytes]:
    """Google Translate TTS (thường cho giọng nữ). Trả về mp3 bytes."""
    if gTTS is None:
        return None
    text = (text or "").strip()
    if not text:
        return None
    try:
        fp = BytesIO()
        gTTS(text=text, lang="vi", tld="com.vn").write_to_fp(fp)
        return fp.getvalue()
    except Exception:
        return None


def _play_mp3_autoplay(mp3_bytes: bytes) -> None:
    """Phát mp3 ngay trong UI (autoplay sau khi bấm nút)."""
    if not mp3_bytes:
        return
    b64 = base64.b64encode(mp3_bytes).decode("ascii")
    html = f"""
      <audio controls autoplay style="width:100%;">
        <source src="data:audio/mpeg;base64,{b64}" type="audio/mpeg">
      </audio>
    """
    st.components.v1.html(html, height=62)


def speak_wiki_vi_from_payload(payload: Dict[str, Any], max_chars: int = 900) -> None:
    """Đọc phần Wikipedia tiếng Việt trong payload bằng giọng Google (ưu tiên gTTS)."""
    intro_vi = (payload.get("wiki_intro_vi") or "").strip()
    title = (payload.get("wiki_title_used") or payload.get("wiki_title_query") or payload.get("common") or payload.get("scientific") or "").strip()

    if not intro_vi:
        intro_vi = f"Không tìm thấy nội dung Wikipedia tiếng Việt cho {title or 'đối tượng này'}."

    text = _clean_wiki_vi_for_tts(intro_vi, max_chars=max_chars)

    mp3 = _gtts_vi_mp3(text)
    if mp3:
        _play_mp3_autoplay(mp3)
        return

    # Fallback: dùng TTS của trình duyệt, ưu tiên voice Google nếu có
    if gTTS is None:
        st.info("Chưa cài gTTS nên đang dùng TTS của trình duyệt (nếu Chrome có voice Google, sẽ ưu tiên).")
    browser_tts_sequence([{"text": text, "lang": "vi-VN", "prefer_google": True}])

    vi_name = common or scientific
    en_name = scientific or common

    if read_vi and vi_name:
        items.append({"text": vi_name, "lang": "vi-VN"})
    if read_en and en_name:
        if not (read_vi and vi_name and en_name.lower() == vi_name.lower()):
            items.append({"text": en_name, "lang": "en-US"})
    return items


# =========================================================
# WIKIPEDIA (VI + EN, ổn định hơn, có Wikidata sitelinks)
# =========================================================
WIKI_OK_TTL_SEC = 30 * 24 * 3600
WIKI_EMPTY_TTL_SEC = 15 * 60

WIKI_LOCK = threading.Lock()
WIKI_SESSION = requests.Session()


def _now() -> float:
    return time.time()


def _cache_get(cache: Dict[str, Any], key: str) -> Optional[Dict[str, Any]]:
    v = cache.get(key)
    if not isinstance(v, dict):
        return None
    try:
        ts = float(v.get("ts", 0.0) or 0.0)
    except Exception:
        return None
    intro = (v.get("intro") or "").strip()
    ttl = WIKI_OK_TTL_SEC if intro else WIKI_EMPTY_TTL_SEC
    if (_now() - ts) <= ttl:
        return v
    return None


def _cache_set(cache: Dict[str, Any], key: str, used: str, intro: str, qid: str = "", err: str = "") -> None:
    cache[key] = {
        "ts": _now(),
        "used": used or "",
        "intro": intro or "",
        "qid": qid or "",
        "error": err or "",
    }


def _wiki_api(lang: str, params: Dict[str, Any], user_agent: str, timeout: float = 10.0) -> Dict[str, Any]:
    url = f"https://{lang}.wikipedia.org/w/api.php"
    headers = {"User-Agent": user_agent}

    last_err: Optional[Exception] = None
    for attempt in range(4):
        try:
            r = WIKI_SESSION.get(url, params=params, headers=headers, timeout=timeout)

            if r.status_code == 429:
                retry_after = r.headers.get("Retry-After")
                sleep_s = float(retry_after) if (retry_after and retry_after.isdigit()) else (0.8 + 0.6 * attempt)
                time.sleep(sleep_s)
                continue

            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep((0.4 + 0.4 * attempt) + random.random() * 0.2)

    raise RuntimeError(f"Wikipedia API failed ({lang}): {last_err}")


def _wiki_fetch_extract_title_qid(lang: str, title: str, user_agent: str) -> Tuple[str, str, str]:
    title = (title or "").strip()
    if not title:
        return "", "", ""

    params = {
        "action": "query",
        "prop": "extracts|pageprops",
        "exintro": 1,
        "explaintext": 1,
        "redirects": 1,
        "format": "json",
        "titles": title,
    }
    data = _wiki_api(lang, params, user_agent=user_agent)
    pages = (data.get("query", {}) or {}).get("pages", {}) or {}
    page = next(iter(pages.values()), {}) if pages else {}

    if page.get("missing") is not None or page.get("pageid") in (-1, None):
        return title, "", ""

    used = (page.get("title") or title).strip()
    extract = (page.get("extract") or "").strip()
    intro = "\n".join(extract.splitlines()[:3]).strip() if extract else ""

    pageprops = page.get("pageprops") or {}
    qid = (pageprops.get("wikibase_item") or "").strip()

    return used, intro, qid


def _wiki_search_best_title(lang: str, query: str, user_agent: str) -> str:
    query = (query or "").strip()
    if not query:
        return ""
    params = {
        "action": "query",
        "list": "search",
        "srsearch": query,
        "srlimit": 1,
        "format": "json",
        "utf8": 1,
    }
    data = _wiki_api(lang, params, user_agent=user_agent)
    hits = ((data.get("query") or {}).get("search") or [])
    return (hits[0].get("title") if hits else "") or ""


def _wikidata_sitelinks(qid: str, user_agent: str) -> Dict[str, str]:
    qid = (qid or "").strip()
    if not qid:
        return {}

    url = "https://www.wikidata.org/w/api.php"
    params = {
        "action": "wbgetentities",
        "ids": qid,
        "props": "sitelinks",
        "format": "json",
    }
    headers = {"User-Agent": user_agent}

    for attempt in range(3):
        try:
            r = WIKI_SESSION.get(url, params=params, headers=headers, timeout=10)
            if r.status_code == 429:
                time.sleep(0.8 + 0.6 * attempt)
                continue
            r.raise_for_status()
            data = r.json()
            ent = (data.get("entities") or {}).get(qid) or {}
            sitelinks = ent.get("sitelinks") or {}

            out: Dict[str, str] = {}
            if "viwiki" in sitelinks:
                out["vi"] = (sitelinks["viwiki"].get("title") or "").strip()
            if "enwiki" in sitelinks:
                out["en"] = (sitelinks["enwiki"].get("title") or "").strip()
            return out
        except Exception:
            time.sleep((0.4 + 0.4 * attempt) + random.random() * 0.2)

    return {}


def wiki_intro_vi_en(query: str, cache: Dict[str, Any], user_agent: str) -> Dict[str, str]:
    query = (query or "").strip()
    if not query:
        return {
            "intro_vi": "",
            "intro_en": "",
            "used_lang": "none",
            "title_used_vi": "",
            "title_used_en": "",
            "title_used_final": "",
            "qid": "",
        }

    key_vi_q = f"vi:{query}".lower()
    key_en_q = f"en:{query}".lower()
    cvi = _cache_get(cache, key_vi_q)
    cen = _cache_get(cache, key_en_q)
    if cvi and cen and (cvi.get("intro") or "").strip() and (cen.get("intro") or "").strip():
        intro_vi = cvi.get("intro", "") or ""
        intro_en = cen.get("intro", "") or ""
        used_vi = cvi.get("used", query) or query
        used_en = cen.get("used", query) or query
        return {
            "intro_vi": intro_vi,
            "intro_en": intro_en,
            "used_lang": "vi,en",
            "title_used_vi": used_vi,
            "title_used_en": used_en,
            "title_used_final": used_vi or used_en or query,
            "qid": (cvi.get("qid") or cen.get("qid") or ""),
        }

    with WIKI_LOCK:
        try:
            qid = ""

            used_vi, intro_vi, qid_vi = _wiki_fetch_extract_title_qid("vi", query, user_agent)
            used_en, intro_en, qid_en = _wiki_fetch_extract_title_qid("en", query, user_agent)
            qid = qid_vi or qid_en

            if not intro_vi:
                best = _wiki_search_best_title("vi", query, user_agent)
                if best and best != query:
                    used_vi, intro_vi, qid_vi2 = _wiki_fetch_extract_title_qid("vi", best, user_agent)
                    qid = qid or qid_vi2

            if not intro_en:
                best = _wiki_search_best_title("en", query, user_agent)
                if best and best != query:
                    used_en, intro_en, qid_en2 = _wiki_fetch_extract_title_qid("en", best, user_agent)
                    qid = qid or qid_en2

            if qid:
                sl = _wikidata_sitelinks(qid, user_agent)
                if sl.get("vi"):
                    used_vi, intro_vi, _ = _wiki_fetch_extract_title_qid("vi", sl["vi"], user_agent)
                if sl.get("en"):
                    used_en, intro_en, _ = _wiki_fetch_extract_title_qid("en", sl["en"], user_agent)

            key_vi_used = f"vi:{(used_vi or query)}".lower()
            key_en_used = f"en:{(used_en or query)}".lower()

            _cache_set(cache, key_vi_used, used_vi or query, intro_vi or "", qid=qid)
            _cache_set(cache, key_en_used, used_en or query, intro_en or "", qid=qid)
            _cache_set(cache, key_vi_q, used_vi or query, intro_vi or "", qid=qid)
            _cache_set(cache, key_en_q, used_en or query, intro_en or "", qid=qid)

            save_json(WIKI_CACHE, cache)

        except Exception as e:
            err = str(e)
            _cache_set(cache, key_vi_q, query, "", qid="", err=err)
            _cache_set(cache, key_en_q, query, "", qid="", err=err)
            save_json(WIKI_CACHE, cache)
            used_vi, used_en, intro_vi, intro_en, qid = query, query, "", "", ""

    used_langs = []
    if (intro_vi or "").strip():
        used_langs.append("vi")
    if (intro_en or "").strip():
        used_langs.append("en")

    title_used_final = used_vi if (intro_vi or "").strip() else (used_en if (intro_en or "").strip() else (used_vi or used_en or query))

    return {
        "intro_vi": intro_vi or "",
        "intro_en": intro_en or "",
        "used_lang": ",".join(used_langs) if used_langs else "none",
        "title_used_vi": used_vi or "",
        "title_used_en": used_en or "",
        "title_used_final": title_used_final or "",
        "qid": qid or "",
    }


# =========================================================
# IDENTIFICATION APIS
# =========================================================
def identify_plantnet(image_jpeg_bytes: bytes, api_key: str, lang: str = "vi") -> Dict[str, Any]:
    api_key = (api_key or "").strip()
    if not api_key:
        return {"ok": False, "error": "Bạn chưa set PLANTNET_API_KEY."}

    url = "https://my-api.plantnet.org/v2/identify/all"
    params = {
        "api-key": api_key,
        "lang": lang,
        "include-related-images": "false",
    }

    files = {"images": ("image.jpg", image_jpeg_bytes, "image/jpeg")}
    data = {"organs": "auto"}

    try:
        r = requests.post(url, params=params, files=files, data=data, timeout=30)
        r.raise_for_status()
        j = r.json()
    except Exception as e:
        return {"ok": False, "error": f"PlantNet error: {e}"}

    results = j.get("results") or []
    best = results[0] if results else {}
    score = float(best.get("score", 0.0) or 0.0)
    species = best.get("species") or {}
    sci_wo_author = (species.get("scientificNameWithoutAuthor") or "").strip()
    sci_full = (species.get("scientificName") or "").strip()
    common_names = species.get("commonNames") or []

    return {
        "ok": True,
        "source": "plantnet",
        "score": score,
        "scientific": sci_wo_author or sci_full or "",
        "common": (common_names[0] if common_names else ""),
        "kingdom_guess": "plant",
        "raw": j,
    }


def identify_insectid(
    image_bytes_any: bytes,
    api_key: str,
    details: str = "url,common_names",
    include_similar_images: bool = False,
) -> Dict[str, Any]:
    api_key = (api_key or "").strip()
    if not api_key:
        return {"ok": False, "error": "Bạn chưa set INSECT_ID_API_KEY."}

    norm_jpeg, norm_info = normalize_for_kindwise(image_bytes_any)
    if not norm_jpeg:
        return {"ok": False, "error": "Không chuẩn hoá được ảnh cho InsectID.", "debug": norm_info}

    url = "https://insect.kindwise.com/api/v1/identification"
    params = {"details": (details or "").strip()}
    base_headers = {"Api-Key": api_key}

    attempts_debug = []

    def try_json_payload(b64_str: str) -> Tuple[bool, Dict[str, Any]]:
        headers = {"Content-Type": "application/json", **base_headers}
        payload = {"images": [b64_str]}
        # ✅ chỉ gửi khi TRUE (KHÔNG gửi false) để tránh lỗi "Unknown modifier"
        if include_similar_images:
            payload["similar_images"] = True

        r = requests.post(url, params=params, headers=headers, json=payload, timeout=KINDWISE_TIMEOUT)
        if 200 <= r.status_code < 300:
            return True, {"json": r.json()}
        return False, {"debug": _resp_debug(r)}

    def try_multipart(jpeg_bytes: bytes) -> Tuple[bool, Dict[str, Any]]:
        files = {"images": ("image.jpg", jpeg_bytes, "image/jpeg")}
        data = {}
        # ✅ chỉ gửi khi TRUE
        if include_similar_images:
            data["similar_images"] = "true"

        r = requests.post(url, params=params, headers=base_headers, files=files, data=data, timeout=KINDWISE_TIMEOUT)
        if 200 <= r.status_code < 300:
            return True, {"json": r.json()}
        return False, {"debug": _resp_debug(r)}

    b64_plain = base64.b64encode(norm_jpeg).decode("ascii")

    ok, out = try_json_payload(b64_plain)
    attempts_debug.append({"attempt": "json_base64_plain", "normalize": norm_info, **out})
    if not ok:
        b64_dataurl = f"data:image/jpeg;base64,{b64_plain}"
        ok2, out2 = try_json_payload(b64_dataurl)
        attempts_debug.append({"attempt": "json_base64_dataurl", "normalize": norm_info, **out2})
        if not ok2:
            ok3, out3 = try_multipart(norm_jpeg)
            attempts_debug.append({"attempt": "multipart_form", "normalize": norm_info, **out3})
            if not ok3:
                return {"ok": False, "error": "InsectID HTTP request failed.", "debug": {"attempts": attempts_debug}}
            j = out3["json"]
        else:
            j = out2["json"]
    else:
        j = out["json"]

    suggestions = j.get("result", {}).get("classification", {}).get("suggestions") or []
    best = suggestions[0] if suggestions else {}
    if not best:
        return {"ok": False, "error": "InsectID: Không có suggestions.", "debug": {"normalize": norm_info, "response": j}}

    score = float(best.get("probability", 0.0) or 0.0)
    sci_name = (best.get("name") or "").strip()
    common_names = best.get("details", {}).get("common_names") or []
    common = (common_names[0] if common_names else "")

    return {
        "ok": True,
        "source": "insectid",
        "score": score,
        "scientific": sci_name,
        "common": common,
        "kingdom_guess": "insect",
        "raw": j,
        "debug": {"normalize": norm_info, "attempts": [x.get("attempt") for x in attempts_debug]},
    }


def pick_best(mode: str, img_jpeg_bytes: bytes, plantnet_lang: str) -> Dict[str, Any]:
    mode = (mode or "auto").lower()
    if mode == "plant":
        return identify_plantnet(img_jpeg_bytes, PLANTNET_API_KEY, lang=plantnet_lang)
    if mode == "insect":
        return identify_insectid(img_jpeg_bytes, INSECT_ID_API_KEY)

    r1 = identify_plantnet(img_jpeg_bytes, PLANTNET_API_KEY, lang=plantnet_lang)
    r2 = identify_insectid(img_jpeg_bytes, INSECT_ID_API_KEY)

    if not r1.get("ok") and r2.get("ok"):
        return r2
    if not r2.get("ok") and r1.get("ok"):
        return r1
    if not r1.get("ok") and not r2.get("ok"):
        # Ưu tiên trả lỗi PlantNet như code cũ, bạn có thể đổi nếu muốn
        return r1

    return r1 if float(r1.get("score", 0.0)) >= float(r2.get("score", 0.0)) else r2


# =========================================================
# RESULT MODEL
# =========================================================
@dataclass
class DetectionResult:
    ts_iso: str
    mode: str
    source: str
    kingdom_guess: str
    scientific: str
    common: str
    score: float
    status: str
    img_path: str
    wiki_title_query: str
    wiki_used_lang: str
    wiki_title_used: str
    wiki_intro_vi: str
    wiki_intro_en: str


def overlay_text(frame_bgr: np.ndarray, text: str) -> np.ndarray:
    out = frame_bgr.copy()
    x, y = 14, 34
    font = cv2.FONT_HERSHEY_SIMPLEX
    scale = 0.80
    thickness = 2
    (tw, th), baseline = cv2.getTextSize(text, font, scale, thickness)
    cv2.rectangle(out, (x - 10, y - th - 10), (x + tw + 10, y + baseline + 10), (0, 0, 0), -1)
    cv2.putText(out, text, (x, y), font, scale, (255, 255, 255), thickness, cv2.LINE_AA)
    return out


# =========================================================
# REALTIME TRANSFORMER
# =========================================================
class DetectorTransformer(VideoTransformerBase):
    def __init__(self):
        self.lock = threading.Lock()
        self.last_call_t = 0.0
        self.last_saved_t = 0.0
        self.last_result: Optional[DetectionResult] = None
        self.last_preview_jpeg: Optional[bytes] = None  # để hiển thị trong popup

        # controlled by UI
        self.enabled = False
        self.interval_sec = 5.0
        self.save_cooldown_sec = 8.0
        self.mode = "auto"
        self.plantnet_lang = "vi"
        self.conf_thresh = 0.55
        self.user_agent = WIKI_USER_AGENT_DEFAULT

    def transform(self, frame):
        img = frame.to_ndarray(format="bgr24")
        raw = img.copy()

        with self.lock:
            lr = self.last_result

        if lr:
            label = lr.scientific or lr.common or "unknown"
            txt = f"{label} ({lr.score:.2f}) [{lr.status}]"
            img = overlay_text(img, txt)

        now = time.time()
        do_call = False
        with self.lock:
            if self.enabled and (now - self.last_call_t) >= self.interval_sec:
                self.last_call_t = now
                do_call = True

        if do_call:
            # Realtime đã là JPEG thật
            ok, buf = cv2.imencode(".jpg", raw, [int(cv2.IMWRITE_JPEG_QUALITY), 90])
            img_bytes = buf.tobytes() if ok else b""
            if not img_bytes:
                return img

            res = pick_best(self.mode, img_bytes, self.plantnet_lang)
            if not res.get("ok"):
                return img

            score = float(res.get("score", 0.0))
            scientific = (res.get("scientific") or "").strip()
            common = (res.get("common") or "").strip()
            source = res.get("source", "")
            kingdom_guess = res.get("kingdom_guess", "")

            status = "KNOWN" if score >= self.conf_thresh else "UNKNOWN"
            label_for_file = scientific or common or "unknown"
            out_dir = COLLECTION_DIR if status == "KNOWN" else UNKNOWN_DIR

            # Wikipedia VI + EN
            cache = load_json(WIKI_CACHE, {})
            wiki_title_query = scientific or common
            wiki_used_lang = "none"
            wiki_title_used = ""
            intro_vi = ""
            intro_en = ""
            if wiki_title_query:
                wk = wiki_intro_vi_en(wiki_title_query, cache, user_agent=self.user_agent)
                intro_vi = wk["intro_vi"]
                intro_en = wk["intro_en"]
                wiki_used_lang = wk["used_lang"]
                wiki_title_used = wk["title_used_final"]

            # save image occasionally
            img_path = ""
            with self.lock:
                self.last_preview_jpeg = img_bytes  # giữ ảnh preview luôn
                if (now - self.last_saved_t) >= self.save_cooldown_sec:
                    self.last_saved_t = now
                    img_path = save_jpeg_bytes(img_bytes, out_dir, label_for_file, score)

            ts_iso = datetime.now().isoformat(timespec="milliseconds")

            if img_path:
                append_log(
                    {
                        "timestamp": ts_iso,
                        "mode": self.mode,
                        "source": source,
                        "kingdom_guess": kingdom_guess,
                        "label_scientific": scientific,
                        "label_common": common,
                        "confidence": score,
                        "status": status,
                        "image_path": img_path,
                        "wiki_title_query": wiki_title_query,
                        "wiki_used_lang": wiki_used_lang,
                        "wiki_title_used": wiki_title_used,
                        "wiki_intro_vi": intro_vi,
                        "wiki_intro_en": intro_en,
                    }
                )

            det = DetectionResult(
                ts_iso=ts_iso,
                mode=self.mode,
                source=source,
                kingdom_guess=kingdom_guess,
                scientific=scientific,
                common=common,
                score=score,
                status=status,
                img_path=img_path,
                wiki_title_query=wiki_title_query or "",
                wiki_used_lang=wiki_used_lang,
                wiki_title_used=wiki_title_used or "",
                wiki_intro_vi=intro_vi or "",
                wiki_intro_en=intro_en or "",
            )

            with self.lock:
                self.last_result = det

        return img


# =========================================================
# POPUP (Dialog / Modal)
# =========================================================
def _render_result_content(payload: Dict[str, Any], read_vi: bool, read_en: bool) -> None:
    """Nội dung bên trong popup."""
    scientific = payload.get("scientific", "") or ""
    common = payload.get("common", "") or ""
    status = payload.get("status", "") or ""
    score = float(payload.get("score") or 0.0)
    mode = payload.get("mode", "") or ""
    source = payload.get("source", "") or ""
    ts = payload.get("ts", "") or ""
    kingdom_guess = payload.get("kingdom_guess", "") or ""
    img_path = payload.get("img_path", "") or ""
    img_bytes = payload.get("img_bytes")

    col1, col2 = st.columns([1, 1])

    with col1:
        if img_bytes:
            st.image(img_bytes, caption="Ảnh dùng để nhận diện", use_container_width=True)
        elif img_path and os.path.exists(img_path):
            st.image(img_path, caption="Ảnh đã lưu", use_container_width=True)
        else:
            st.info("Không có ảnh preview (realtime có thể chưa tới lượt lưu).")

        st.markdown("#### Thông tin")
        kind = "known" if status.upper() == "KNOWN" else "unknown" if status.upper() == "UNKNOWN" else "info"
        st.markdown(badge_html(f"Status: {status}", kind), unsafe_allow_html=True)
        st.write(f"**Score:** `{score:.3f}`")
        st.progress(min(1.0, max(0.0, score)))

        st.write(f"**Scientific:** {scientific or '-'}")
        st.write(f"**Common:** {common or '-'}")
        st.write(f"**Mode:** `{mode}`  |  **Source:** `{source}`")
        if kingdom_guess:
            st.write(f"**Kingdom guess:** `{kingdom_guess}`")
        if ts:
            st.write(f"**Time:** {ts}")
        if img_path:
            st.caption(f"Saved: `{img_path}`")

        st.divider()
        cA, cB = st.columns(2)
        with cA:
            if st.button("🔊 Đọc Wikipedia (VI)", key=f"{st.session_state.get('popup_keyprefix', 'popup')}_speak"):
                if read_vi:
                    speak_wiki_vi_from_payload(payload)
                else:
                    st.info("Bạn đang tắt TTS Wikipedia (VI) ở sidebar.")
        with cB:
            if st.button("❌ Đóng", key=f"{st.session_state.get('popup_keyprefix', 'popup')}_close"):
                st.session_state["popup_open"] = False
                st.session_state["popup_payload"] = None
                st.rerun()

    with col2:
        st.markdown("#### Wikipedia")
        wiki_title_query = payload.get("wiki_title_query", "") or ""
        wiki_used_lang = payload.get("wiki_used_lang", "") or ""
        wiki_title_used = payload.get("wiki_title_used", "") or ""
        intro_vi = payload.get("wiki_intro_vi", "") or ""
        intro_en = payload.get("wiki_intro_en", "") or ""

        st.caption(f"Query: `{wiki_title_query}` • Used: `{wiki_used_lang}` • Title: `{wiki_title_used}`")

        tab_vi, tab_en = st.tabs(["🇻🇳 VI", "🇺🇸 EN"])
        with tab_vi:
            st.text_area("Giới thiệu (VI)", value=intro_vi, height=220)
        with tab_en:
            st.text_area("Intro (EN)", value=intro_en, height=220)


def open_result_popup(title: str, payload: Dict[str, Any], read_vi: bool, read_en: bool) -> None:
    """
    Mở popup theo version Streamlit:
    - Ưu tiên st.dialog / st.experimental_dialog
    - Nếu không có, fallback hiển thị 'giả popup' bằng container.
    """
    if hasattr(st, "dialog"):
        @st.dialog(title, width="large")
        def _dlg():
            _render_result_content(payload, read_vi=read_vi, read_en=read_en)

        _dlg()
        return

    if hasattr(st, "experimental_dialog"):
        @st.experimental_dialog(title, width="large")
        def _dlg2():
            _render_result_content(payload, read_vi=read_vi, read_en=read_en)

        _dlg2()
        return

    # Fallback (không đúng nghĩa popup, nhưng vẫn "nổi" và dễ nhìn)
    st.warning("Streamlit phiên bản hiện tại không hỗ trợ dialog/modal. Mình hiển thị kết quả trong khung bên dưới.")
    card_open()
    st.markdown(f"### {title}")
    _render_result_content(payload, read_vi=read_vi, read_en=read_en)
    card_close()


def set_popup(payload: Dict[str, Any], title: str) -> None:
    st.session_state["popup_open"] = True
    st.session_state["popup_title"] = title
    st.session_state["popup_payload"] = payload
    # prefix ổn định cho widget keys trong popup (tránh random mỗi rerun)
    ts = (payload or {}).get("ts", "") or datetime.now().isoformat(timespec="milliseconds")
    st.session_state["popup_keyprefix"] = f"{ts}_{int(time.time()*1000)}"


# =========================================================
# STREAMLIT UI
# =========================================================
def main():
    st.set_page_config(page_title=APP_NAME, page_icon="🪲", layout="wide")

    ensure_dirs()
    init_log()
    apply_cute_theme()

    # Session state
    st.session_state.setdefault("last_rt", None)             # dict lưu kết quả realtime gần nhất
    st.session_state.setdefault("last_spoken_rt_ts", "")     # để auto đọc 1 lần mỗi kết quả mới
    st.session_state.setdefault("last_popup_rt_ts", "")      # để auto mở popup 1 lần mỗi kết quả mới
    st.session_state.setdefault("last_upload", None)         # dict lưu kết quả upload gần nhất
    st.session_state.setdefault("last_popup_upload_ts", "")
    st.session_state.setdefault("popup_open", False)
    st.session_state.setdefault("popup_title", "Kết quả nhận dạng")
    st.session_state.setdefault("popup_payload", None)
    st.session_state.setdefault("popup_keyprefix", "")

    # Hero header
    st.markdown(
        """
        <div class="cute-hero">
          <h1>🪲🌿 CuteID</h1>
          <p>Nhận diện côn trùng & thực vật (NO BBOX). Upload hoặc Realtime. Kết quả hiển thị trong popup, có TTS đọc Wikipedia (VI).</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.markdown("## ⚙️ Cấu hình")
        card_open()
        mode = st.selectbox("Chế độ nhận diện", ["auto", "plant", "insect"], index=0, help="Auto sẽ gọi cả 2 API (tốn quota).")
        conf_thresh = st.slider("Ngưỡng tin cậy (KNOWN)", 0.0, 1.0, 0.55, 0.01)
        interval_sec = st.slider("Chu kỳ gọi API (realtime)", 1, 30, 5, 1)
        save_cd = st.slider("Cooldown lưu ảnh (realtime)", 1, 60, 8, 1)
        plantnet_lang = st.selectbox("PlantNet language", ["en", "vi", "fr", "es"], index=0)
        user_agent = st.text_input("Wikipedia User-Agent", value=WIKI_USER_AGENT_DEFAULT)
        card_close()

        st.markdown("## 🔊 TTS")
        card_open()
        read_aloud = st.checkbox("Tự đọc Wiki (VI) khi có kết quả mới", value=True)
        read_vi = st.checkbox("Đọc Wikipedia tiếng Việt (giọng Google nữ)", value=True)
        read_en = st.checkbox("Đọc EN (không dùng)", value=False, disabled=True)
        if gTTS is None:
            st.caption("Cài giọng Google nữ: `pip install gTTS` (nếu chưa cài).")
        card_close()

        st.markdown("## 🪄 Popup")
        card_open()
        auto_popup_rt = st.checkbox("Realtime: tự mở popup khi có kết quả mới", value=False)
        auto_popup_upload = st.checkbox("Upload: tự mở popup sau khi nhận diện", value=True)
        card_close()

        st.markdown("## 🔑 API key")
        card_open()
        if not PLANTNET_API_KEY:
            st.warning("Chưa có PLANTNET_API_KEY (env).")
        else:
            st.markdown(badge_html("PlantNet: OK", "known"), unsafe_allow_html=True)

        if not INSECT_ID_API_KEY:
            st.warning("Chưa có INSECT_ID_API_KEY (env).")
        else:
            st.markdown(badge_html("InsectID: OK", "known"), unsafe_allow_html=True)

        st.caption("Tip: dùng env để giữ an toàn. Nếu thiếu key, app vẫn chạy UI nhưng sẽ báo lỗi khi nhận diện.")
        card_close()

        if st_autorefresh is None:
            st.caption("Cài thêm để realtime mượt hơn: `pip install streamlit-autorefresh`")

    tab1, tab2 = st.tabs(["🎥 Realtime", "📷 Upload / Camera"])

    # ========================
    # REALTIME TAB
    # ========================
    with tab1:
        st.markdown("### 🎥 Realtime (WebRTC)")
        col_cam, col_ctrl = st.columns([2, 1])

        with col_cam:
            card_open()
            webrtc_ctx = webrtc_streamer(
                key="realtime",
                mode=WebRtcMode.SENDRECV,
                video_transformer_factory=DetectorTransformer,
                media_stream_constraints={"video": True, "audio": False},
                async_processing=True,
            )
            card_close()

        with col_ctrl:
            card_open()
            st.markdown("#### Điều khiển")
            enabled = st.toggle("Bật nhận diện realtime", value=False) if hasattr(st, "toggle") else st.checkbox("Bật nhận diện realtime", value=False)

            if enabled and st_autorefresh is not None:
                st_autorefresh(interval=500, key="rt_autorefresh")  # 0.5s

            lr = None
            preview = None
            if webrtc_ctx and webrtc_ctx.video_transformer:
                tr: DetectorTransformer = webrtc_ctx.video_transformer
                tr.enabled = enabled
                tr.interval_sec = float(interval_sec)
                tr.save_cooldown_sec = float(save_cd)
                tr.mode = mode
                tr.conf_thresh = float(conf_thresh)
                tr.plantnet_lang = plantnet_lang
                tr.user_agent = user_agent or WIKI_USER_AGENT_DEFAULT

                with tr.lock:
                    lr = tr.last_result
                    preview = tr.last_preview_jpeg

            # Store last realtime result into session_state
            if lr:
                st.session_state["last_rt"] = {
                    "ts": lr.ts_iso,
                    "scientific": lr.scientific,
                    "common": lr.common,
                    "status": lr.status,
                    "score": lr.score,
                    "mode": lr.mode,
                    "source": lr.source,
                    "kingdom_guess": lr.kingdom_guess,
                    "img_path": lr.img_path,
                    "img_bytes": preview,
                    "wiki_title_query": lr.wiki_title_query,
                    "wiki_used_lang": lr.wiki_used_lang,
                    "wiki_title_used": lr.wiki_title_used,
                    "wiki_intro_vi": lr.wiki_intro_vi,
                    "wiki_intro_en": lr.wiki_intro_en,
                }

            last = st.session_state.get("last_rt")

            st.divider()
            st.markdown("#### Kết quả gần nhất")
            if not last:
                st.info("Chưa có kết quả. Hãy bật realtime và đưa đối tượng vào khung hình.")
            else:
                kind = "known" if (last.get("status", "").upper() == "KNOWN") else "unknown"
                st.markdown(badge_html(f"{last.get('status','')}", kind), unsafe_allow_html=True)
                st.write(f"**{last.get('common') or last.get('scientific') or 'unknown'}**")
                st.caption(f"Score: {float(last.get('score') or 0.0):.3f} • Source: {last.get('source','')} • Mode: {last.get('mode','')}")

                cA, cB = st.columns(2)
                with cA:
                    if st.button("✨ Xem popup", use_container_width=True):
                        set_popup(last, "✨ Kết quả nhận dạng (Realtime)")
                with cB:
                    if st.button("🔁 Đọc Wiki (VI)", use_container_width=True):
                        if read_vi:
                            speak_wiki_vi_from_payload(last)
                        else:
                            st.info("Bạn đang tắt TTS Wikipedia (VI) ở sidebar.")

                # Auto TTS once per new detection
                if enabled and read_aloud:
                    if st.session_state.get("last_spoken_rt_ts", "") != last.get("ts", ""):
                        if read_vi:
                            speak_wiki_vi_from_payload(last)
                            st.session_state["last_spoken_rt_ts"] = last.get("ts", "")

                # Auto popup once per new detection (optional)
                if enabled and auto_popup_rt and last.get("ts"):
                    if st.session_state.get("last_popup_rt_ts", "") != last.get("ts", ""):
                        set_popup(last, "✨ Kết quả nhận dạng (Realtime)")
                        st.session_state["last_popup_rt_ts"] = last.get("ts", "")

            card_close()

    # ========================
    # UPLOAD TAB
    # ========================
    with tab2:
        st.markdown("### 📷 Upload / Camera (ảnh tĩnh)")
        col_in, col_out = st.columns([1, 2])

        with col_in:
            card_open()
            up = st.file_uploader("Upload ảnh (jpg/png)", type=["jpg", "jpeg", "png"])
            cam = st.camera_input("Hoặc chụp ảnh")
            card_close()

        raw_bytes = None
        if up is not None:
            raw_bytes = up.getvalue()
        elif cam is not None:
            raw_bytes = cam.getvalue()

        with col_out:
            card_open()
            if raw_bytes is None:
                st.info("Chọn ảnh hoặc chụp ảnh để nhận diện.")
                card_close()
            else:
                frame_bgr = decode_image_to_bgr(raw_bytes)
                if frame_bgr is None:
                    st.error("Không đọc được ảnh.")
                    card_close()
                else:
                    st.image(cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2RGB), caption="Ảnh đầu vào", use_container_width=True)

                    # IMPORTANT FIX: luôn ép JPEG thật
                    img_jpeg_bytes = encode_bgr_to_jpeg_bytes(frame_bgr, quality=92)
                    if not img_jpeg_bytes:
                        st.error("Không encode được ảnh sang JPEG.")
                        card_close()
                    else:
                        st.divider()
                        c1, c2, c3 = st.columns([1, 1, 1])
                        with c1:
                            do_identify = st.button("✨ Nhận diện", type="primary", use_container_width=True)
                        with c2:
                            if st.button("🧹 Xoá kết quả", use_container_width=True):
                                st.session_state["last_upload"] = None
                                st.session_state["popup_open"] = False
                                st.session_state["popup_payload"] = None
                                st.rerun()
                        with c3:
                            if st.button("🔁 Đọc Wiki (VI) (Upload)", use_container_width=True):
                                lastu = st.session_state.get("last_upload")
                                if lastu:
                                    if read_vi:
                                        speak_wiki_vi_from_payload(lastu)
                                    else:
                                        st.info("Bạn đang tắt TTS Wikipedia (VI) ở sidebar.")
                                else:
                                    st.info("Chưa có kết quả upload để đọc lại.")

                        if do_identify:
                            if not PLANTNET_API_KEY and not INSECT_ID_API_KEY:
                                st.error("Bạn chưa set API key (PLANTNET_API_KEY / INSECT_ID_API_KEY).")
                            else:
                                with st.spinner("Đang gọi API..."):
                                    res = pick_best(mode, img_jpeg_bytes, plantnet_lang)

                                if not res.get("ok"):
                                    st.error(res.get("error", "Lỗi không xác định"))
                                    if res.get("debug"):
                                        with st.expander("Chi tiết lỗi (debug)"):
                                            st.json(res.get("debug"))
                                else:
                                    score = float(res.get("score", 0.0))
                                    scientific = (res.get("scientific") or "").strip()
                                    common = (res.get("common") or "").strip()
                                    source = res.get("source", "")
                                    kingdom_guess = res.get("kingdom_guess", "")

                                    status = "KNOWN" if score >= float(conf_thresh) else "UNKNOWN"
                                    label_for_file = scientific or common or "unknown"
                                    out_dir = COLLECTION_DIR if status == "KNOWN" else UNKNOWN_DIR
                                    img_path = save_jpeg_bytes(img_jpeg_bytes, out_dir, label_for_file, score)

                                    # Wikipedia VI + EN
                                    cache = load_json(WIKI_CACHE, {})
                                    wiki_title_query = scientific or common
                                    wiki_used_lang = "none"
                                    wiki_title_used = ""
                                    intro_vi = ""
                                    intro_en = ""
                                    if wiki_title_query:
                                        wk = wiki_intro_vi_en(wiki_title_query, cache, user_agent=user_agent or WIKI_USER_AGENT_DEFAULT)
                                        intro_vi = wk["intro_vi"]
                                        intro_en = wk["intro_en"]
                                        wiki_used_lang = wk["used_lang"]
                                        wiki_title_used = wk["title_used_final"]

                                    ts_iso = datetime.now().isoformat(timespec="milliseconds")

                                    append_log(
                                        {
                                            "timestamp": ts_iso,
                                            "mode": mode,
                                            "source": source,
                                            "kingdom_guess": kingdom_guess,
                                            "label_scientific": scientific,
                                            "label_common": common,
                                            "confidence": score,
                                            "status": status,
                                            "image_path": img_path,
                                            "wiki_title_query": wiki_title_query,
                                            "wiki_used_lang": wiki_used_lang,
                                            "wiki_title_used": wiki_title_used,
                                            "wiki_intro_vi": intro_vi,
                                            "wiki_intro_en": intro_en,
                                        }
                                    )

                                    payload = {
                                        "ts": ts_iso,
                                        "scientific": scientific,
                                        "common": common,
                                        "status": status,
                                        "score": score,
                                        "mode": mode,
                                        "source": source,
                                        "kingdom_guess": kingdom_guess,
                                        "img_path": img_path,
                                        "img_bytes": img_jpeg_bytes,
                                        "wiki_title_query": wiki_title_query,
                                        "wiki_used_lang": wiki_used_lang,
                                        "wiki_title_used": wiki_title_used,
                                        "wiki_intro_vi": intro_vi,
                                        "wiki_intro_en": intro_en,
                                    }
                                    st.session_state["last_upload"] = payload

                                    st.success("Xong! ✅")
                                    try:
                                        st.toast("Đã nhận diện xong ✨", icon="✅")
                                    except Exception:
                                        pass

                                    # Auto TTS (đọc Wikipedia VI)
                                    if read_aloud and read_vi:
                                        speak_wiki_vi_from_payload(payload)

                                    # Auto popup (default ON)
                                    if auto_popup_upload:
                                        set_popup(payload, "✨ Kết quả nhận dạng (Upload)")
                                        st.session_state["last_popup_upload_ts"] = ts_iso

                        # Small inline summary + open popup
                        lastu = st.session_state.get("last_upload")
                        st.divider()
                        st.markdown("#### Kết quả upload gần nhất")
                        if not lastu:
                            st.caption("Chưa có kết quả.")
                        else:
                            kind = "known" if (lastu.get("status", "").upper() == "KNOWN") else "unknown"
                            st.markdown(badge_html(f"{lastu.get('status','')}", kind), unsafe_allow_html=True)
                            st.write(f"**{lastu.get('common') or lastu.get('scientific') or 'unknown'}**")
                            st.caption(f"Score: {float(lastu.get('score') or 0.0):.3f} • Source: {lastu.get('source','')}")

                            if st.button("✨ Xem popup (Upload)", use_container_width=True):
                                set_popup(lastu, "✨ Kết quả nhận dạng (Upload)")

                        card_close()

    # Global popup render (cuối file để luôn render sau khi state set)
    if st.session_state.get("popup_open") and st.session_state.get("popup_payload"):
        open_result_popup(
            st.session_state.get("popup_title", "Kết quả nhận dạng"),
            st.session_state["popup_payload"],
            read_vi=read_vi,
            read_en=read_en,
        )

    st.markdown("---")
    st.caption("Gợi ý: Mode 'auto' gọi cả 2 API (tốn quota). Biết chắc là cây/côn trùng thì chọn đúng để tiết kiệm.")


if __name__ == "__main__":
    main()
